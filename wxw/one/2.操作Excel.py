from openpyxl import load_workbook

wb = load_workbook(filename='F:\\sublime\\python\\doc\\wxw.xlsx')

print('文件名称：',wb.sheetnames)

sheet=wb['用户信息']

print(sheet)
print('表的内容：',sheet.cell(2,2).value)






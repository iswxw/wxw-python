from openpyxl import Workbook

wb = Workbook()  # 创建文件对象

# 获取第一个sheet
ws = wb.active

# 将数据写入到指定的单元格
ws['A1'] = '编号'  # 写入数据
ws['B1'] = '姓名' # 写入中文
ws['C1'] = '年龄'

ws.append([1,'魏小伟',18])  # 写入多个单元格

# 修改当前表的标题
ws.title='用户信息'

# 查看工作的表
print('所有的工作表：',wb.sheetnames)

# 遍历所有表
for sheet in wb:
	print('工作表：',sheet.title)


# 保存为 .xlsx
wb.save('F:\\sublime\\python\\doc\\wxw.xlsx')



# 获取单元格的值
for row in ws.values:
   for value in row:
     print(value,end=',')
   print('')


# 打印为只读
print('===========')
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
	print(row)